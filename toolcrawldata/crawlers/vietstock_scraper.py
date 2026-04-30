"""
Vietstock News Scraper
Tải danh sách bài viết + nội dung đầy đủ từng bài, xuất ra Excel.

Nguồn: https://vietstock.vn/chu-de/1-8/tat-ca.htm
Chạy:  python3 vietstock_scraper.py
"""

import sys
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, date, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import time

# Import shared Supabase writer từ thư
#  mục toolcrawldata/
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from supabase_writer import get_client, upsert_batch

# ── Cấu hình ──────────────────────────────────────────────────────────────────
LOOKBACK_HOURS = 24                # Chỉ lấy bài trong 24h gần nhất
MAX_ARTICLES   = 250               # Tự dừng nếu vượt quá số này
MAX_PAGES      = 50                # Giới hạn trang an toàn
ITEMS_PER_PAGE = 15

OUTPUT_FILE = str(Path(__file__).parent / f"vietstock_news_{date.today().strftime('%m%Y')}.xlsx")
WORKERS       = 5                  # Số luồng song song khi tải nội dung
CONTENT_DELAY = 0.3                # Giây chờ giữa mỗi request nội dung

API_URL  = "https://vietstock.vn/_Partials/GetStockNewsByMarketPaging"
SITE_URL = "https://vietstock.vn"

API_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Content-Type": "application/json; charset=UTF-8",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://vietstock.vn/chu-de/1-8/tat-ca.htm",
    "Origin": "https://vietstock.vn",
}

WEB_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept-Language": "vi-VN,vi;q=0.9",
    "Referer": "https://vietstock.vn/",
}

_lock = threading.Lock()
_done = 0


# ── Bước 1: Tải danh sách bài ─────────────────────────────────────────────────

def parse_publish_time(ts_str: str) -> datetime | None:
    try:
        ms = int(ts_str.replace("/Date(", "").replace(")/", ""))
        return datetime.utcfromtimestamp(ms / 1000) + timedelta(hours=7)
    except Exception:
        return None


def fetch_article_list(session: requests.Session, page: int) -> dict | None:
    try:
        resp = session.post(
            API_URL,
            json={"item": ITEMS_PER_PAGE, "martket": "1", "row": page},
            headers=API_HEADERS,
            timeout=20,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        print(f"  [!] Lỗi API trang {page}: {e}")
        return None


def scrape_article_list(lookback_hours: int = LOOKBACK_HOURS) -> list[dict]:
    """
    Lấy bài trong `lookback_hours` giờ gần nhất.
    Tự dừng nếu đạt MAX_ARTICLES hoặc gặp bài cũ hơn ngưỡng.
    """
    cutoff   = datetime.now() - timedelta(hours=lookback_hours)
    session  = requests.Session()
    articles = []

    print(f"  Vietstock: lấy bài từ {cutoff.strftime('%d/%m/%Y %H:%M')} ({lookback_hours}h)")

    for page in range(1, MAX_PAGES + 1):
        data = fetch_article_list(session, page)

        if not data or data.get("Code") != 200 or not data.get("Data"):
            print(f"  Vietstock: hết dữ liệu ở trang {page}, dừng.")
            break

        stop = False
        for art in data["Data"]:
            pub_time = parse_publish_time(art.get("PublishTime", ""))
            if not pub_time:
                continue

            # Dừng khi gặp bài cũ hơn ngưỡng 24h
            if pub_time < cutoff:
                stop = True
                break

            url = art.get("URL", "")
            if url and not url.startswith("http"):
                url = SITE_URL + url

            articles.append({
                "Tiêu đề":       art.get("Title", "").strip(),
                "Nội dung":      "",
                "Ngày đăng":     pub_time.strftime("%d/%m/%Y %H:%M"),
                "Mã CK":         art.get("StockCode", ""),
                "Tác giả":       art.get("By", ""),
                "Giá đóng cửa":  art.get("ClosePrice"),
                "Thay đổi giá":  art.get("Change"),
                "% Thay đổi":    art.get("PerChange"),
                "Link bài viết": url,
                "Link tài chính": art.get("FinanceURL", ""),
                "_ts":           pub_time,
            })

            # Giới hạn tối đa 250 bài
            if len(articles) >= MAX_ARTICLES:
                print(f"  Vietstock: đạt giới hạn {MAX_ARTICLES} bài, dừng.")
                stop = True
                break

        print(f"  Vietstock trang {page}: +{len(data['Data'])} | tổng {len(articles)}")
        if stop:
            break
        time.sleep(0.8)

    print(f"  Vietstock: {len(articles)} bài trong {lookback_hours}h\n")
    return articles


# ── Bước 2: Tải nội dung từng bài ────────────────────────────────────────────

def fetch_content(row_idx: int, url: str) -> tuple[int, str]:
    global _done
    if not url or not url.startswith("http"):
        return row_idx, ""
    try:
        resp = requests.get(url, headers=WEB_HEADERS, timeout=15)
        resp.raise_for_status()
        resp.encoding = "utf-8"
        soup = BeautifulSoup(resp.text, "html.parser")

        article = (soup.select_one("#page-content")
                   or soup.select_one(".article-content"))
        if not article:
            return row_idx, ""

        for tag in article.select(
            ".article-sharing, .rightdetail, .scroll-content-sub, "
            "script, style, .ads, [class*=ads], .related, "
            ".tags, .author-info, h1, .meta, .dateNewBlock"
        ):
            tag.decompose()

        paragraphs = article.select("p")
        texts = [p.get_text(separator=" ", strip=True) for p in paragraphs]
        if texts and len(texts[0]) < 150 and "." not in texts[0]:
            texts = texts[1:]

        content = "\n\n".join(t for t in texts if t)
        time.sleep(CONTENT_DELAY)

    except Exception as e:
        content = f"[Lỗi: {e}]"

    with _lock:
        _done += 1
        if _done % 100 == 0 or _done <= 3:
            print(f"  [{_done}] {url[:70]}")

    return row_idx, content


def enrich_content(articles: list[dict]) -> list[dict]:
    global _done
    _done = 0
    total = len(articles)
    urls = [a["Link bài viết"] for a in articles]

    print(f"{'='*55}")
    print(f"  BƯỚC 2: Tải nội dung {total} bài ({WORKERS} luồng song song)")
    print(f"  Ước tính: ~{int(total * CONTENT_DELAY / WORKERS / 60) + 1} phút")
    print(f"{'='*55}")

    start = time.time()
    contents = [""] * total

    with ThreadPoolExecutor(max_workers=WORKERS) as executor:
        futures = {executor.submit(fetch_content, i, url): i for i, url in enumerate(urls)}
        for future in as_completed(futures):
            idx, content = future.result()
            contents[idx] = content

    elapsed = time.time() - start
    print(f"\n  -> Hoàn thành trong {elapsed:.0f}s\n")

    for i, content in enumerate(contents):
        articles[i]["Nội dung"] = content

    return articles


# ── Bước 3: Upsert Supabase ───────────────────────────────────────────────────

def upsert_news_to_supabase(articles: list[dict]) -> None:
    """Upsert danh sách bài viết vào bảng market_news."""
    if not articles:
        return

    records = []
    seen_urls = set()
    for a in articles:
        url = a.get("Link bài viết") or None
        # Normalize fili.vn → vietstock.vn (cùng domain, tránh trùng bài)
        if url and 'fili.vn' in url:
            url = url.replace('fili.vn', 'vietstock.vn').replace('http://', 'https://')
        if url and url in seen_urls:
            continue
        if url:
            seen_urls.add(url)
        pub_ts = a.get("_ts")
        records.append({
            "symbol":           a.get("Mã CK") or None,
            "title":            a.get("Tiêu đề", ""),
            "content":          a.get("Nội dung") or None,
            "author":           a.get("Tác giả") or None,
            "source":           "vietstock",
            "close_price":      a.get("Giá đóng cửa") or None,
            "price_change":     a.get("Thay đổi giá") or None,
            "price_change_pct": a.get("% Thay đổi") or None,
            "article_url":      url,
            "finance_url":      a.get("Link tài chính") or None,
            "published_at":     pub_ts.isoformat() if pub_ts else None,
        })

    print(f"{'='*55}")
    print(f"  BƯỚC 3b: Upsert {len(records)} bài lên Supabase")
    print(f"{'='*55}")

    try:
        sb = get_client()
        n = upsert_batch(sb, "market_news", records, "article_url")
        print(f"  ✓ Supabase market_news: {n} rows upserted\n")
    except Exception as e:
        print(f"  ✗ Supabase lỗi: {e}\n")


# ── Bước 4: Xuất Excel ────────────────────────────────────────────────────────

def export_to_excel(articles: list[dict], output_path: str):  # noqa: C901
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"{'='*55}")
    print(f"  BƯỚC 3: Xuất Excel")
    print(f"{'='*55}")

    df = pd.DataFrame(articles)
    df = df.sort_values("_ts", ascending=False).drop(columns=["_ts"]).reset_index(drop=True)
    df.index += 1

    cols = [
        "Tiêu đề", "Nội dung", "Ngày đăng", "Mã CK", "Tác giả",
        "Giá đóng cửa", "Thay đổi giá", "% Thay đổi",
        "Link bài viết", "Link tài chính",
    ]
    df = df[[c for c in cols if c in df.columns]]

    col_widths = {
        "STT": 5, "Tiêu đề": 50, "Nội dung": 80, "Ngày đăng": 16,
        "Mã CK": 8, "Tác giả": 12, "Giá đóng cửa": 14,
        "Thay đổi giá": 13, "% Thay đổi": 11,
        "Link bài viết": 50, "Link tài chính": 45,
    }
    headers = ["STT"] + list(df.columns)
    content_col_idx = headers.index("Nội dung") + 1 if "Nội dung" in headers else None

    thin = Side(style="thin", color="D0D7E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="EEF2FB")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Tin tức Vietstock", index=True, index_label="STT")
        ws = writer.sheets["Tin tức Vietstock"]

        # Header row
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 28

        # Column widths
        for i, col_name in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 15)

        # Data rows
        num_cols = {"Giá đóng cửa", "Thay đổi giá", "% Thay đổi"}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for cell in row:
                col_name = headers[cell.column - 1] if cell.column <= len(headers) else ""
                cell.fill = fill
                cell.border = border
                wrap = (content_col_idx is not None and cell.column == content_col_idx)
                cell.alignment = Alignment(vertical="top", wrap_text=wrap)
                if col_name in num_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                    if col_name != "Giá đóng cửa":
                        cell.font = Font(color="009900" if cell.value > 0 else "CC0000")

        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions

        # Summary sheet
        ws2 = writer.book.create_sheet("Thống kê")
        ws2["A1"] = f"THỐNG KÊ TIN TỨC VIETSTOCK — {LOOKBACK_HOURS}h GẦN NHẤT"
        ws2["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws2.merge_cells("A1:D1")
        ws2["A1"].alignment = Alignment(horizontal="center")

        bold = Font(bold=True)
        summary_rows = [
            ("Tổng số bài viết:", len(df)),
            ("Từ ngày:", df["Ngày đăng"].iloc[-1] if len(df) else ""),
            ("Đến ngày:", df["Ngày đăng"].iloc[0] if len(df) else ""),
            ("Có nội dung đầy đủ:",
             df["Nội dung"].apply(lambda x: bool(x) and not str(x).startswith("[Lỗi")).sum()),
            ("Thời gian xuất:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Nguồn:", "https://vietstock.vn/chu-de/1-8/tat-ca.htm"),
        ]
        for i, (label, value) in enumerate(summary_rows, start=3):
            ws2.cell(row=i, column=1, value=label).font = bold
            ws2.cell(row=i, column=2, value=value)

        if "Mã CK" in df.columns:
            ws2["A10"] = "TOP MÃ CHỨNG KHOÁN XUẤT HIỆN NHIỀU NHẤT"
            ws2["A10"].font = Font(bold=True, size=11, color="1F3864")
            ws2["A11"] = "Mã CK"
            ws2["B11"] = "Số bài"
            for cell in [ws2["A11"], ws2["B11"]]:
                cell.fill = PatternFill("solid", fgColor="1F3864")
                cell.font = Font(bold=True, color="FFFFFF")
            top = df[df["Mã CK"].notna() & (df["Mã CK"] != "")]["Mã CK"].value_counts().head(20)
            for i, (code, cnt) in enumerate(top.items(), start=12):
                ws2.cell(row=i, column=1, value=code)
                ws2.cell(row=i, column=2, value=cnt)

        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 15

    print(f"\n  File: {output_path}")
    print(f"  {len(df)} bài viết đã lưu\n")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\nVIETSTOCK SCRAPER — bắt đầu lúc {datetime.now().strftime('%H:%M:%S')}\n")

    # Bước 1: danh sách bài
    articles = scrape_article_list()
    if not articles:
        print("[!] Không lấy được danh sách bài.")
        sys.exit(1)

    # Bước 2: nội dung đầy đủ
    articles = enrich_content(articles)

    # Bước 3a: upsert lên Supabase
    upsert_news_to_supabase(articles)

    # Bước 3b: xuất Excel (giữ lại làm backup local)
    export_to_excel(articles, OUTPUT_FILE)

    print(f"HOÀN THÀNH lúc {datetime.now().strftime('%H:%M:%S')}")
